import os
import re
from pathlib import Path
from datetime import datetime
import PyPDF2
from pdfminer.high_level import extract_text
from langdetect import detect, DetectorFactory
from openpyxl import Workbook
from openpyxl.styles import Font

# Ensure consistent language detection results
DetectorFactory.seed = 0

def extract_pdf_metadata(pdf_path):
    """Extract metadata from PDF file"""
    metadata = {
        'title': 'Unknown',
        'author': 'Unknown',
        'creator': 'Unknown',
        'producer': 'Unknown',
        'creation_date': 'Unknown',
        'pages': 0,
        'language': 'Unknown'
    }
    
    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            
            # Extract basic metadata
            info = pdf_reader.metadata
            if info:
                if info.title:
                    metadata['title'] = info.title
                if info.author:
                    metadata['author'] = info.author
                if info.creator:
                    metadata['creator'] = info.creator
                if info.producer:
                    metadata['producer'] = info.producer
                if info.creation_date:
                    try:
                        date_str = info.creation_date.decode() if isinstance(info.creation_date, bytes) else str(info.creation_date)
                        year_match = re.search(r'(\d{4})', date_str)
                        if year_match:
                            metadata['creation_date'] = year_match.group(1)
                    except:
                        metadata['creation_date'] = 'Unknown'
            
            # Get page count
            metadata['pages'] = len(pdf_reader.pages)
            
            # Try to extract text from first few pages for language detection
            text_sample = ""
            try:
                for page_num in range(min(3, len(pdf_reader.pages))):
                    page = pdf_reader.pages[page_num]
                    text_sample += page.extract_text()
                
                if text_sample.strip():
                    try:
                        metadata['language'] = detect(text_sample)
                    except:
                        metadata['language'] = 'Unknown'
            except:
                pass
                
    except Exception as e:
        print(f"Error processing {pdf_path}: {str(e)}")
    
    return metadata

def extract_text_metadata(pdf_path):
    """Extract additional metadata from PDF text using pdfminer"""
    text_metadata = {
        'publisher': 'Unknown',
        'year': 'Unknown'
    }
    
    try:
        # Extract text from first few pages
        text = extract_text(pdf_path, maxpages=5)
        
        # Look for publisher patterns
        publisher_patterns = [
            r'published by\s+([^\n,.]+)',
            r'publisher\s*:\s*([^\n,.]+)',
            r'©\s*\d{4}\s*([^\n,.]+)',
            r'copyright\s*\d{4}\s*([^\n,.]+)'
        ]
        
        for pattern in publisher_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                text_metadata['publisher'] = match.group(1).strip()
                break
        
        # Look for year patterns
        year_patterns = [
            r'copyright\s*(\d{4})',
            r'©\s*(\d{4})',
            r'published\s*in\s*(\d{4})',
            r'first published\s*(\d{4})'
        ]
        
        for pattern in year_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                text_metadata['year'] = match.group(1).strip()
                break
                
    except Exception as e:
        print(f"Error extracting text from {pdf_path}: {str(e)}")
    
    return text_metadata

def process_pdf_directory(directory_path):
    """Process all PDF files in the given directory"""
    catalog_data = []
    
    pdf_files = list(Path(directory_path).glob("*.pdf"))
    
    if not pdf_files:
        print(f"No PDF files found in {directory_path}")
        return catalog_data
    
    print(f"Found {len(pdf_files)} PDF files to process...")
    
    for pdf_path in pdf_files:
        print(f"Processing: {pdf_path.name}")
        
        # Extract metadata using PyPDF2
        pdf_metadata = extract_pdf_metadata(pdf_path)
        
        # Extract additional metadata from text
        text_metadata = extract_text_metadata(pdf_path)
        
        # Combine metadata
        book_info = {
            'Book Title': pdf_metadata['title'],
            'Author': pdf_metadata['author'],
            'Editor': pdf_metadata['creator'],
            'Year of Publishing': text_metadata['year'] if text_metadata['year'] != 'Unknown' else pdf_metadata['creation_date'],
            'Publisher': text_metadata['publisher'] if text_metadata['publisher'] != 'Unknown' else pdf_metadata['producer'],
            'Language': pdf_metadata['language'],
            'Number of Pages': pdf_metadata['pages'],
            'Format': 'PDF'
        }
        
        catalog_data.append(book_info)
    
    return catalog_data

def create_excel_catalog(catalog_data, output_path):
    """Create Excel catalog from extracted data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Book Catalog"
    
    # Define column headers
    headers = [
        'Book Title', 'Author', 'Editor', 'Year of Publishing', 
        'Publisher', 'Language', 'Number of Pages', 'Format'
    ]
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True)
    
    # Write data
    for row_num, book_info in enumerate(catalog_data, 2):
        for col_num, header in enumerate(headers, 1):
            value = book_info.get(header, 'Unknown')
            ws.cell(row=row_num, column=col_num, value=value)
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    wb.save(output_path)
    print(f"Excel catalog created: {output_path}")

def main():
    """Main function to run the PDF catalog generator"""
    # Get directory path from user
    directory_path = input("Enter the path to the directory containing PDF files: ").strip()
    
    if not os.path.isdir(directory_path):
        print("Error: The specified path is not a valid directory.")
        return
    
    # Process PDF files
    catalog_data = process_pdf_directory(directory_path)
    
    if not catalog_data:
        print("No data extracted from PDF files.")
        return
    
    # Create output filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_filename = f"book_catalog_{timestamp}.xlsx"
    output_path = os.path.join(directory_path, output_filename)
    
    # Generate Excel catalog
    create_excel_catalog(catalog_data, output_path)
    
    print(f"Catalog generation complete. Processed {len(catalog_data)} books.")

if __name__ == "__main__":
    main()